import glob
import os
import hashlib
from openpyxl import Workbook
import PySimpleGUI as sg
from pygost.gost34112012256 import GOST34112012256

ResList = list[(str, str)]

def get_hash_f_sha256(path: str) ->str:
    with open(path, "rb") as f:
        bytes = f.read()
        return hashlib.sha256(bytes).hexdigest()

def get_gost34112012_256(path: str) -> str:
    with open(path, "rb") as f:
        bytes = f.read()
        hex_str = GOST34112012256(bytes).hexdigest()
        # print("".join([hex_str[i : i + 2] for i in range(0, len(hex_str), 2)][::-1]))
        return "".join([hex_str[i : i + 2] for i in range(0, len(hex_str), 2)][::-1]).upper()

def get_data(root: str) -> ResList:
    buff: ResList = []

    for filename in glob.iglob(root + '**/**', recursive=True):
        if os.path.isfile(filename):
            #hash = get_hash_f_sha256(filename)
            hash = get_gost34112012_256(filename)
            # print(f"{filename} - {hash}")
            buff.append((filename, hash))
    return buff

def print2excel(buff: ResList, fname: str):
    wb = Workbook()
    sheet = wb.active
    sheet.cell(1, 1, "Файл")
    sheet.cell(1, 2, "Расширение")
    sheet.cell(1, 3, "Хэш")
    sheet.cell(1, 4, "Путь")

    for (index, (fn, hash)) in enumerate(buff, start = 2):
        (head, tail) = os.path.split(fn)
        (_, exten) = os.path.splitext(tail)
        sheet.cell(index, 1, tail)
        sheet.cell(index, 2, exten[1:].upper())
        sheet.cell(index, 3, hash)
        sheet.cell(index, 4, head.replace("/", "\\"))
    wb.save(fname)

# def run(root : str):
#     print(root)
#     buff: ResList = get_data(root)
#
#     # print2excel(buff, "log.xlsx")
#     for (index, (fn, hsh)) in enumerate(buff, start = 1):
#         print(f"{index}. {fn} - {hsh}")

layout = [
            [
                sg.Text("Выберете папку"),
                sg.In(size=(40, 1), enable_events=True, key="-SRC_DIR-"),
                sg.FolderBrowse(button_text="Выбрать")
            ],
            [
                sg.Text("Сохранить в...  "),
                sg.In(size=(40, 1), enable_events=True, key="-DST_FILE-"),
                sg.FileSaveAs(button_text="Выбрать", file_types=(("Книга Excel", "*.xlsx"),))
            ],
            # [
            #     sg.Checkbox('SHA256', key="SHA256")
            # ],
            [
                sg.Button("Выполнить", enable_events=True, key="-RUN-"),
                sg.Button("Выйти", enable_events=True, key="Exit")
            ]
]
# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    # root_dir = os.path.join("D:\\", "Workdir", "444")
    # run(root_dir)
    root = os.path.dirname(__file__)
    images = os.path.join(root, "images/GREEN.ico")

    window = sg.Window("Файлы - ХЕШ ГОСТ Р 34.11-2012 (256)", layout, icon=images)
    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED or event in ("Exit", "Cancel"):
            break
        if event == "-RUN-":
            if values["-SRC_DIR-"] == "":
                sg.popup_error('Выберете папку с файлами!')
            elif values["-DST_FILE-"] == "":
                sg.popup_error('Выберете куда сохранить!')
            else:
                buff: ResList = get_data(values["-SRC_DIR-"])
                print2excel(buff, values["-DST_FILE-"])
                sg.popup_ok('Выполнено!')
    window.close()
